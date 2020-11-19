import mysql.connector
from mysql.connector import Error
import pandas as pd
import openpyxl
import os

path = 'D:\pt'

def identifiercheck(project_name):
    if '山语墅' in project_name:
        identifier = 1
    elif '西区供水' in project_name:
        identifier = 2
    elif '避暑文化产业园' in project_name:
        identifier = 3
    else:
        identifier = 0
    return identifier

def specimentypetrim(typename):
    m = ''
    for i in typename:
        if i != '-':
            m += i
        else:
            return m
    return('Unvalid')

def create_server_connection(host_name, user_name, user_password):
    connection = None
    try:
        connection = mysql.connector.connect(
            host=host_name,
            user=user_name,
            passwd=user_password
        )
        print("MySQL Database connection successful")
    except Error as err:
        print(f"Error: '{err}'")

    return connection

pw = "jingwei" # IMPORTANT! Put your MySQL Terminal password here.
db = "ztinfo11" # This is the name of the database we will create in the next step - call it whatever you like.

connection = create_server_connection("localhost", "root", pw)

def create_database(connection, query):
    cursor = connection.cursor()
    try:
        cursor.execute(query)
        print("Database created successfully")
    except Error as err:
        print(f"Error: '{err}'")

create_database_query = 'CREATE DATABASE ' + db 
create_database(connection, create_database_query)

def create_db_connection(host_name, user_name, user_password, db_name):
    connection = None
    try:
        connection = mysql.connector.connect(
            host=host_name,
            user=user_name,
            passwd=user_password,
            database=db_name
        )
        print("MySQL Database connection successful")
    except Error as err:
        print(f"Error: '{err}'")

    return connection

def execute_query(connection, query):
    cursor = connection.cursor()
    try:
        cursor.execute(query)
        connection.commit()
        cursor.close()
        print("Query successful")
    except Error as err:
        print(f"Error: '{err}'")

def read_query(connection, query):
    cursor = connection.cursor()
    result = None
    try:
        cursor.execute(query)
        result = cursor.fetchall()
        return result
    except Error as err:
        print(f"Error: '{err}'")

def create_database():
    create_project_table = """
    CREATE TABLE project (
      project_id INT PRIMARY KEY,
      project_name VARCHAR(80) NOT NULL,
      request_date DATE NOT NULL,
      status VARCHAR(20) NOT NULL,
      price FLOAT(10) NOT NULL,
      specimen_type VARCHAR(20) NOT NULL,
      company_name VARCHAR(60),
      location VARCHAR(60),
      company_identifier INT
      );
     """


    create_company_table = """
    CREATE TABLE company (
      company_id INT PRIMARY KEY,
      company_name VARCHAR(40),
      deposit float(15),
      amount_payed float(20)
    );
     """

    create_collection_history_table = """
    CREATE TABLE collection_history (
      collection_id INT PRIMARY KEY,
      collector_identifier INT NOT NULL,
      collected_num INT NOT NULL,
      collected_value float(20) NOT NULL,
      collected_date DATE NOT NULL
    );
     """

    connection = create_db_connection("localhost", "root", pw, db) # Connect to the Database
    execute_query(connection, create_project_table) 
    execute_query(connection, create_company_table) 
    execute_query(connection, create_collection_history_table) 

def read_data():
    connection = create_db_connection("localhost", "root", pw, db)
    files=os.listdir(path)
    exfiles=[f for f in files if f.endswith('.xlsx')]
    for exfile in exfiles:
        book = openpyxl.load_workbook(path+'/'+exfile)
        sheet = book.active
        for row in range(3,sheet.max_row+1):
            buffer = []
            for infoindex in [1,2,4,8,9,10,11,15]: #类 号 状态 日期 委托方 工程部位 工程名称 价格
                buffer.append(sheet.cell(row,infoindex).value)
            specimen_type = specimentypetrim(buffer[0])
            project_id = buffer[1]
            status = buffer[2]
            request_date = buffer[3]
            company_name = buffer[4]
            location = buffer[5]
            project_name = buffer[6]
            price = buffer[7]
            company_identifier = identifiercheck(project_name)
            if specimen_type !='12': #in ['MC','MQ','SKY']
                k = f"{project_id},'{project_name}','{request_date}','{status}',{price},'{specimen_type}','{company_name}','{location}',{company_identifier}"
                pop_project_1 = """
                INSERT INTO project(project_id, project_name, request_date, status, price, specimen_type, company_name, location, company_identifier) VALUES
                (
                """
                pop_project_2 = """
                )
                """
                pop_project = pop_project_1 + k + pop_project_2
                execute_query(connection,pop_project)
        q3 = """
        SELECT *
        FROM project
        """
        #connection = create_db_connection("localhost", "root", pw, db)
        results = read_query(connection, q3)

        from_db = []

        for result in results:
            result = list(result)
            from_db.append(result)
    
        print(from_db)
def count():

    from_db_l = []
    accumulation = 0.0
    type_list = []
    type_count = []

    connection = create_db_connection("localhost", "root", pw, db)

    project_idinput = input('Type project_id here: ')
    project_id_list = project_idinput.split()
    
    for project_id in project_id_list:
        
        if '-' in project_id:
            
            project_id_maxmin = project_id.split('-')
            sq = f'''
            SELECT price,specimen_type FROM project
            WHERE project_id>={project_id_maxmin[0]} and project_id<={project_id_maxmin[1]}
            '''
            results=read_query(connection,sq) 
            from_db_l.append(results)
        
        else:
            
            sq = f'''
            SELECT price,specimen_type FROM project
            WHERE project_id={project_id}
            '''
            results=read_query(connection,sq) 
            from_db_l.append(results)
    print(from_db_l)
    if not(isinstance(from_db_l[0][0],float)):
        for from_db in from_db_l:
            for i in range(0,len(from_db)):
                if from_db[i][0] != 0.0:
                    if from_db[i][1] in type_list:
                        index = type_list.index(from_db[i][1])
                        type_count[index] = type_count[index] + 1
                        accumulation = accumulation + from_db[i][0]
                    else:
                        type_list.append(from_db[i][1])
                        type_count.append(1)
                        accumulation = accumulation + from_db[i][0]
    else:
        print('t2')
        from_db = from_db_l
        for i in range(0,len(from_db)):
            print('t3')
            if from_db[i][0] != 0.0:
                print('t4')
                if from_db[i][1] in type_list:
                    print('t5')
                    index = type_list.index(from_db[i][1])
                    print('t6')
                    type_count[index] = type_count[index] + 1
                    print('t7')
                    accumulation = accumulation + from_db[i][0]
                else:
                    print('t8')
                    type_list.append(from_db[i][1])
                    print('t9')
                    type_count.append(1)
                    print('t10')
                    accumulation = accumulation + from_db[i][0]
    print(type_list,type_count,accumulation)



while __name__=='__main__':
    command = input('Type your cmd here: ')
    if command == 'setup':
        create_database()
        read_data()
    elif command == 'count':
        count()
